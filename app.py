import streamlit as st
import pandas as pd
import streamlit.components.v1 as components
import os
import json

# Check for openpyxl availability (not needed anymore, but kept for compatibility)
try:
    import openpyxl  # type: ignore
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ==========================================
# 1. CONFIGURATION
# ==========================================
st.set_page_config(page_title="ORO Logic Capturer", page_icon="🚦", layout="wide")

st.markdown("""
<style>
    .header-style {font-size: 24px; font-weight: bold; margin-bottom: 10px;}
    .green-lane {border-left: 5px solid #22c55e; background-color: #f0fdf4; padding: 20px; border-radius: 8px;}
    .red-lane {border-left: 5px solid #ef4444; background-color: #fef2f2; padding: 20px; border-radius: 8px;}
    .stDataFrame {border: 1px solid #e2e8f0; border-radius: 5px;}
</style>
""", unsafe_allow_html=True)

# ==========================================
# 2. HELPER FUNCTIONS: DATA LOADING
# ==========================================

def load_geo_from_df(df):
    """Convert Geo DataFrame to GEO_HIERARCHY dictionary"""
    hierarchy = {}
    if df is not None and not df.empty:
        for _, row in df.iterrows():
            # Используем strip() для удаления пробелов и проверяем на пустые строки
            region = str(row.get('Region', '')).strip() if pd.notna(row.get('Region', '')) else ''
            drbu = str(row.get('DRBU', '')).strip() if pd.notna(row.get('DRBU', '')) else ''
            end_market = str(row.get('End Market', '')).strip() if pd.notna(row.get('End Market', '')) else ''
            
            # Загружаем все строки, даже если некоторые поля пустые (но не все)
            if region or drbu or end_market:
                if not region:
                    region = 'Unknown'
                if not drbu:
                    drbu = 'Unknown'
                if not end_market:
                    end_market = 'Unknown'
                
                if region not in hierarchy:
                    hierarchy[region] = {}
                if drbu not in hierarchy[region]:
                    hierarchy[region][drbu] = []
                if end_market not in hierarchy[region][drbu]:
                    hierarchy[region][drbu].append(end_market)
    return hierarchy

def load_cat_from_df(df):
    """Convert Categories DataFrame to CAT_HIERARCHY dictionary"""
    hierarchy = {}
    if df is not None and not df.empty:
        for _, row in df.iterrows():
            l1 = row.get('L1', '')
            l2 = row.get('L2', '')
            l3 = row.get('L3', '')
            l4 = row.get('L4', '')
            
            if l1 and l2 and l3 and l4:
                if l1 not in hierarchy:
                    hierarchy[l1] = {}
                if l2 not in hierarchy[l1]:
                    hierarchy[l1][l2] = {}
                if l3 not in hierarchy[l1][l2]:
                    hierarchy[l1][l2][l3] = []
                if l4 not in hierarchy[l1][l2][l3]:
                    hierarchy[l1][l2][l3].append(l4)
    return hierarchy

# ==========================================
# 3. DATA: LOAD FROM FILE OR USE DEFAULTS
# ==========================================

# Initialize session state for data
if 'geo_df' not in st.session_state:
    st.session_state.geo_df = None
if 'cat_df' not in st.session_state:
    st.session_state.cat_df = None
if 'geo_hierarchy' not in st.session_state:
    st.session_state.geo_hierarchy = None
if 'cat_hierarchy' not in st.session_state:
    st.session_state.cat_hierarchy = None

# Default hierarchies (fallback)
DEFAULT_GEO_HIERARCHY = {
  "AME": {
    "WESTERN EUROPE": [
      "Belgium", 
      "Denmark", 
      "France", 
      "Greece", 
      "Ireland", 
      "Malta", 
      "Norway", 
      "Spain", 
      "Sweden", 
      "UNITED KINGDOM", 
      "Netherlands", 
      "Finland", 
      "Luxembourg", 
      "Portugal", 
      "Cyprus"
    ]
  },
  "APMEA": {
    "APMEA SOUTH": [
      "Australia", 
      "Indonesia", 
      "Malaysia", 
      "Papua New Guinea", 
      "Samoa", 
      "Singapore", 
      "Solomon Islands", 
      "Vietnam", 
      "Fiji", 
      "New Zealand", 
      "Cambodia", 
      "Philippines", 
      "China"
    ]
  },
  "USA": {
    "USA": ["USA"]
  }
}

DEFAULT_CAT_HIERARCHY = {
    "Marketing": {
        "Marketing Prof Svc": {
            "Advertising Services": ["Media services"],
            "Creative agency fees": ["Creative agency fees"],
            "Market research": ["Market research-customised", "Market research-syndicated"],
            "Marketing & Trade Event": ["1-2-1 Activation, Brand Ambassadors & Hostesses", "Marketing Event Management", "Sponsorship"],
            "PR Services": ["Partnership", "Printing Cylinders", "PR Agency"],
            "Trade Marketing Services": ["D2C-Social Selling", "Loyalty & Incentive Programmes", "Other Trade Marketing Services"]
        },
        "Marketing POSM": {
            "POSM Services": ["Cigarette Vending Machines Purchase & Lease/Services", "Marketing Print", "Merchandising Services", "Permanent POSM", "POSM Leasing Services", "Promotional merchandise", "Semi-Permanent POSM"]
        }
    },
    "Operations": {
        "Production": {
            "Production Services": ["Manufacturing support services", "Production"],
            "Spare Parts": ["Spare Parts"],
            "After Sales": ["After Sales"]
        },
        "OSS": {
            "Material Handling/Storage Machinery": ["General Machinery", "Material Handling Equipment FLTS", "Warehouse Material (pallets or other consumables)", "Workshop supplies & consumables"],
            "Packaging Materials and Supply": ["Leaf Packaging materials & supplies", "Tobacco Case C48", "Warehouse Packaging Materials"],
            "Quality Control": ["Factory Quality Control and Service", "Quality Control"]
        },
        "Agricultural Inputs": {
            "Agrochemicals (Herbicides, Insecticides, etc.)": ["Agrochemicals (Herbicides, Insecticides, etc.)"],
            "Fertilizers (NPK, Soluble, etc.)": ["Fertilizers (NPK, Soluble, etc.)"],
            "Other Agricultural Inputs (Supplies, Services)": ["Other Agricultural Inputs (Supplies, Services)"]
        }
    },
    "Corporate": {
        "Facilities": {
            "Facilities Services": ["Archiving", "Catering Services and Supplies", "Food & Beverages", "Industrial Cleaning Services", "Integrated Facilities Management", "Landscaping, Roads and Grounds, Snow removal", "Office Cleaning Equipment and Supplies", "Other Facilities", "Plants & Flowers", "Staff Transportation", "Statutory Compliance & Inspections", "Vending Purchase/Lease/Maintenance"],
            "Building": ["Building-Consultancy & Project management", "Building Construction", "Building equipment and installation", "Building Maintenance and Repair"],
            "Corporate Real Estate": ["Property Lease", "Property Purchase or Sale"],
            "Pest control": ["Pest control products & services"],
            "Security Services": ["Security Services and Supplies", "Security Technology & Services"],
            "Uniform": ["Uniform Services and Management"],
            "Utilities": ["Electricity", "Fuels", "Gas", "Utilities other", "Water"],
            "Waste": ["Waste Management Services"]
        },
        "Prof Svc": {
            "Consultancy": ["Consultancy"],
            "Finance Services": ["Banking and investment", "Group Company Auditors"],
            "Legal Services": ["Legal Services", "Legal Services Other", "Litigation", "Patents or Trade Mark"],
            "Other Audits (local), Recovery Audits, Accounting": ["Other Audits (local), Recovery Audits, Accounting"],
            "Translation, Information, Testing, Inspection etc": ["Translation, Information, Testing, Inspection etc"]
        },
        "HR Svc": {
            "HR Professional services": ["HR Consultancy", "Outplacement", "Recruitment", "Training & education"],
            "Reward": ["Benefits & Employee Assistance", "External Payroll Services", "Health & Life Insurance", "Healthcare Services", "HR Compensation and Benefits Surveys", "Pension investments"],
            "Relocation": ["Expats-Schools, House", "Relocation Services"],
            "Talent": ["Temporary Labour and outsourcing", "Temporary Labour IT"]
        },
        "Office Services and supplies": {
            "Office Services and supplies": ["Books-Journals & subcriptions", "Office Equipment", "Office Furniture", "Office supplies", "Printing and Reproduction Services"]
        },
        "Travel Management": {
            "Travel Management": ["Air travel", "Other Travel Expense (Visa, Rail, Sea)", "Taxi-Bus-Car hire"],
            "Hotel-Restaurant & Meeting": ["Hotel", "Restaurant-Bar Expenses", "Seminars-Conference-Meetings"]
        },
        "Vehicle Hire & Purchase": {
            "Vehicle Hire & Purchase": ["Vehicle Lease (long term)", "Vehicle Purchase", "Vehicle rental (short term)"],
            "Other Vehicle Costs": ["Fuel", "Telematics", "Vehicle maintenance/Fleet management", "Vehicle Other Insurance/Tax/Parking"]
        },
        "Insurance": {
            "Insurance": ["Building & Content Insurance", "Insurance Others"]
        },
        "Politics & Civic Affairs": {
            "Politics & Civic Affairs": ["Charities", "Membership fees or Tobacco chambers or unions", "Politics & Civic Affairs"]
        },
        "Other Agency costs": {
            "Other Agency costs": ["Other Agency costs"]
        },
        "Other Travel Expense": {
            "Other Travel Expense": ["Other Travel Expense"]
        }
    },
    "IDT": {
        "IT Infrastructure": {
            "IT Infrastructure": ["Hosting, Public Cloud, Datacentres Infrastructure"],
            "Hardware": ["Computing/Desktop/Laptops/Handheld", "IT Equipment and accessories", "IT Hardware Maintenance", "Servers and Server Equipment"],
            "Networks Hardware": ["Audio and Video Hardware", "IT Networks Infrastructure"],
            "Networks Services": ["WAN & LAN Services"]
        },
        "IT Services": {
            "IT Services": ["IT Services-End User Computing", "IT Services-End User Printing"],
            "IT Consultancy": ["IT Consultancy"],
            "Managed Professional Services": ["Managed Professional Services"]
        },
        "Software & Application": {
            "Software & Application": ["Soft & App Develop Corporate Functions", "Soft & App Develop Econnected Devices", "Soft & App Develop Enterprise Platforms", "Soft & App Develop Marketing D2C", "Soft & App Develop Marketing Trade", "Soft & App Develop Operations", "Soft & App Develop Testing Q&A", "Software License", "Software Support"]
        },
        "Digital Services": {
            "Digital Services": ["CRM Services (Non-DBS)", "Info/Careline/Live Chat/Call Centre", "Other Digital Services", "Search Engine Optimisation (SEO)", "Social Media Management"]
        },
        "Cyber Security": {
            "Cyber Security": ["Cyber Security"]
        },
        "Voice, Communication & Mobile Services": {
            "Voice, Communication & Mobile Services": ["Voice & Mobile Communication Services"]
        }
    },
    "R&D": {
        "Laboratory Supply": {
            "Laboratory Supply": ["Laboratory Consumables", "Laboratory Equipment & Supplies"]
        },
        "Scientific Services": {
            "Scientific Services": ["Analytical", "Clinical Studies", "R&D Consultancy", "Research Services"]
        },
        "EH&S Equipment and Services": {
            "EH&S Equipment and Services": ["Agricultural PPEs (Farmers Protection)", "Safety Equipment & PPEs (Shoes, Gloves, etc.)"]
        },
        "ESG": {
            "ESG": ["ESG Afforestation", "ESG Carbon offsets", "ESG IREC GoO/ESG Renewable energy certificates", "ESG Solar panels"]
        },
        "Equipment": {
            "Equipment": ["Production Machinery"]
        }
    }
}

# ==========================================
# 4. SIDEBAR: SCOPE SELECTION
# ==========================================
with st.sidebar:
    # Initialize default data if not loaded
    if st.session_state.geo_hierarchy is None:
        st.session_state.geo_hierarchy = DEFAULT_GEO_HIERARCHY
    if st.session_state.cat_hierarchy is None:
        st.session_state.cat_hierarchy = DEFAULT_CAT_HIERARCHY
    
    # File uploader removed - using default data only
    
    # Get current DataFrame for geography
    geo_df_current = st.session_state.geo_df
    
    # If no DataFrame, create from default hierarchy
    if geo_df_current is None or geo_df_current.empty:
        # Convert DEFAULT_GEO_HIERARCHY to DataFrame for consistency
        default_geo_data = []
        for region, clusters in DEFAULT_GEO_HIERARCHY.items():
            for cluster, markets in clusters.items():
                for market in markets:
                    default_geo_data.append({
                        'Region': region,
                        'DRBU': cluster,
                        'End Market': market,
                        'Company Code': ''  # Empty value by default
                    })
        geo_df_current = pd.DataFrame(default_geo_data)
        st.session_state.geo_df = geo_df_current
    
    # Determine cluster column name (can be DRBU or Cluster)
    # Support both variants for compatibility
    cluster_col = None
    if 'DRBU' in geo_df_current.columns:
        cluster_col = 'DRBU'
    elif 'Cluster' in geo_df_current.columns:
        cluster_col = 'Cluster'
    else:
        # If neither column exists, try to find similar
        possible_cols = [col for col in geo_df_current.columns if 'cluster' in col.lower() or 'drbu' in col.lower()]
        if possible_cols:
            cluster_col = possible_cols[0]
        else:
            st.error("❌ Missing 'DRBU' or 'Cluster' column in data")
            cluster_col = 'DRBU'  # Fallback to avoid errors
    
    st.divider()
    st.header("1. Scope Definition")
    
    # --- Geography Dropdowns (Cascading) ---
    st.subheader("🌍 Geography")
    
    if geo_df_current is not None and not geo_df_current.empty:
        # 1. Select Region
        regions = sorted(geo_df_current['Region'].unique())
        if not regions:
            st.warning("No region data available")
            region = "N/A"
            cluster = "N/A"
            selected_markets = []
            business_user_markets = []
            company_code = "N/A"
        else:
            region = st.selectbox("Region", regions, key="geo_region")
            
            # 2. Filter Cluster data based on selected Region
            filtered_clusters_df = geo_df_current[geo_df_current['Region'] == region]
            filtered_clusters = sorted(filtered_clusters_df[cluster_col].unique())
            
            if not filtered_clusters:
                st.warning(f"No clusters available for region {region}")
                cluster = "N/A"
                selected_markets = []
                business_user_markets = []
                company_code = "N/A"
            else:
                cluster = st.selectbox("Cluster / DRBU", filtered_clusters, key="geo_cluster")
                
                # 3. Filter Market data based on selected Cluster
                filtered_markets_df = filtered_clusters_df[filtered_clusters_df[cluster_col] == cluster]
                # Remove dropna() to avoid losing data, filter only empty strings
                filtered_markets_df = filtered_markets_df[filtered_markets_df['End Market'].notna()]
                filtered_markets_df = filtered_markets_df[filtered_markets_df['End Market'].astype(str).str.strip() != '']
                filtered_markets = sorted(filtered_markets_df['End Market'].unique().tolist())
                
                if not filtered_markets:
                    st.warning(f"No markets available for cluster {cluster}")
                    selected_markets = []
                    business_user_markets = []
                    company_code = "N/A"
                else:
                    # Multiple selection for End Markets with "Select All" button
                    col_market1, col_market2 = st.columns([3, 1])
                    with col_market1:
                        selected_markets = st.multiselect(
                            "End Market (select one or multiple)", 
                            filtered_markets, 
                            key="geo_market_multiselect",
                            help="Select one or multiple End Markets"
                        )
                    with col_market2:
                        if st.button("Select All", key="select_all_markets", use_container_width=True):
                            selected_markets = filtered_markets
                            st.session_state.geo_market_multiselect = filtered_markets
                            st.rerun()
                        if st.button("Clear All", key="clear_all_markets", use_container_width=True):
                            selected_markets = []
                            st.session_state.geo_market_multiselect = []
                            st.rerun()
                    
                    # Display selected markets
                    if selected_markets:
                        st.caption(f"Selected markets: {len(selected_markets)}")
                        # Show compact list of selected
                        if len(selected_markets) <= 5:
                            st.caption(f"Selected: {', '.join(selected_markets)}")
                        else:
                            st.caption(f"Selected: {', '.join(selected_markets[:5])} and {len(selected_markets) - 5} more")
                    
                    # 4. Business User End Market (multiple selection)
                    st.markdown("---")
                    st.subheader("👤 Business User End Market")
                    col_bu1, col_bu2 = st.columns([3, 1])
                    with col_bu1:
                        business_user_markets = st.multiselect(
                            "Business User End Market (select one or multiple)",
                            filtered_markets,
                            key="business_user_markets",
                            help="Select End Markets for business users"
                        )
                    with col_bu2:
                        if st.button("Select All", key="select_all_bu_markets", use_container_width=True):
                            business_user_markets = filtered_markets
                            st.session_state.business_user_markets = filtered_markets
                            st.rerun()
                        if st.button("Clear All", key="clear_all_bu_markets", use_container_width=True):
                            business_user_markets = []
                            st.session_state.business_user_markets = []
                            st.rerun()
                    
                    if business_user_markets:
                        st.caption(f"Selected business user markets: {len(business_user_markets)}")
                        if len(business_user_markets) <= 5:
                            st.caption(f"Selected: {', '.join(business_user_markets)}")
                        else:
                            st.caption(f"Selected: {', '.join(business_user_markets[:5])} and {len(business_user_markets) - 5} more")
                    
                    # 5. Company Code (if available)
                    if 'Company Code' in geo_df_current.columns:
                        # Filter Company Codes for all selected End Markets
                        if selected_markets:
                            filtered_company_codes_df = filtered_markets_df[filtered_markets_df['End Market'].isin(selected_markets)]
                            filtered_company_codes_df = filtered_company_codes_df[filtered_company_codes_df['Company Code'].notna()]
                            filtered_company_codes_df = filtered_company_codes_df[filtered_company_codes_df['Company Code'].astype(str).str.strip() != '']
                            company_codes = sorted(filtered_company_codes_df['Company Code'].unique().tolist())
                            
                            if company_codes:
                                company_code = st.selectbox("Company Code", company_codes, key="geo_company_code")
                            else:
                                company_code = st.text_input("Company Code (enter manually)", key="geo_company_code_manual", placeholder="e.g., UK001")
                        else:
                            company_code = st.text_input("Company Code (enter manually)", key="geo_company_code_manual", placeholder="Please select End Market first")
                    else:
                        company_code = st.text_input("Company Code", key="geo_company_code_manual", placeholder="e.g., UK001")
    else:
        st.warning("No geography data available")
        region = "N/A"
        cluster = "N/A"
        selected_markets = []
        business_user_markets = []
        company_code = "N/A"
    
    st.divider()
    
    # --- Category Dropdowns (Cascading) ---
    st.subheader("🗂 Category")
    
    # Get current DataFrame for categories
    cat_df_current = st.session_state.cat_df
    
    # If no DataFrame, create from default hierarchy
    if cat_df_current is None or cat_df_current.empty:
        # Convert DEFAULT_CAT_HIERARCHY to DataFrame for consistency
        default_cat_data = []
        for l1, l2_dict in DEFAULT_CAT_HIERARCHY.items():
            for l2, l3_dict in l2_dict.items():
                for l3, l4_list in l3_dict.items():
                    for l4 in l4_list:
                        default_cat_data.append({
                            'L1': l1,
                            'L2': l2,
                            'L3': l3,
                            'L4': l4
                        })
        cat_df_current = pd.DataFrame(default_cat_data)
        st.session_state.cat_df = cat_df_current
    
    if cat_df_current is not None and not cat_df_current.empty:
        # 1. Select L1 Category (multiple selection)
        l1_options = sorted(cat_df_current['L1'].unique())
        if not l1_options:
            st.warning("No L1 category data available")
            selected_l1 = []
            selected_l2 = []
            selected_l3 = []
            selected_l4 = []
            full_cat_path = "N/A > N/A > N/A > N/A"
        else:
            col_l1_1, col_l1_2 = st.columns([3, 1])
            with col_l1_1:
                selected_l1 = st.multiselect("L1 Category (select one or multiple)", l1_options, key="cat_l1_multiselect")
            with col_l1_2:
                if st.button("Select All", key="select_all_l1", use_container_width=True):
                    selected_l1 = l1_options
                    st.session_state.cat_l1_multiselect = l1_options
                    st.rerun()
                if st.button("Clear All", key="clear_all_l1", use_container_width=True):
                    selected_l1 = []
                    st.session_state.cat_l1_multiselect = []
                    st.rerun()
            
            # 2. Filter L2 data based on selected L1(s)
            if selected_l1:
                filtered_l2_df = cat_df_current[cat_df_current['L1'].isin(selected_l1)]
                filtered_l2 = sorted(filtered_l2_df['L2'].unique())
                
                if not filtered_l2:
                    st.warning(f"No L2 categories available for selected L1")
                    selected_l2 = []
                    selected_l3 = []
                    selected_l4 = []
                    full_cat_path = " > ".join(selected_l1) + " > N/A > N/A > N/A"
                else:
                    col_l2_1, col_l2_2 = st.columns([3, 1])
                    with col_l2_1:
                        selected_l2 = st.multiselect("L2 Category (select one or multiple)", filtered_l2, key="cat_l2_multiselect")
                    with col_l2_2:
                        if st.button("Select All", key="select_all_l2", use_container_width=True):
                            selected_l2 = filtered_l2
                            st.session_state.cat_l2_multiselect = filtered_l2
                            st.rerun()
                        if st.button("Clear All", key="clear_all_l2", use_container_width=True):
                            selected_l2 = []
                            st.session_state.cat_l2_multiselect = []
                            st.rerun()
                    
                    # 3. Filter L3 data based on selected L2(s)
                    if selected_l2:
                        filtered_l3_df = filtered_l2_df[filtered_l2_df['L2'].isin(selected_l2)]
                        filtered_l3 = sorted(filtered_l3_df['L3'].unique())
                        
                        if not filtered_l3:
                            st.warning(f"No L3 categories available for selected L2")
                            selected_l3 = []
                            selected_l4 = []
                            full_cat_path = " > ".join(selected_l1) + " > " + " > ".join(selected_l2) + " > N/A > N/A"
                        else:
                            col_l3_1, col_l3_2 = st.columns([3, 1])
                            with col_l3_1:
                                selected_l3 = st.multiselect("L3 Category (select one or multiple)", filtered_l3, key="cat_l3_multiselect")
                            with col_l3_2:
                                if st.button("Select All", key="select_all_l3", use_container_width=True):
                                    selected_l3 = filtered_l3
                                    st.session_state.cat_l3_multiselect = filtered_l3
                                    st.rerun()
                                if st.button("Clear All", key="clear_all_l3", use_container_width=True):
                                    selected_l3 = []
                                    st.session_state.cat_l3_multiselect = []
                                    st.rerun()
                            
                            # 4. Filter L4 data based on selected L3(s)
                            if selected_l3:
                                filtered_l4_df = filtered_l3_df[filtered_l3_df['L3'].isin(selected_l3)]
                                filtered_l4 = sorted(filtered_l4_df['L4'].unique())
                                
                                if not filtered_l4:
                                    st.warning(f"No L4 categories available for selected L3")
                                    selected_l4 = []
                                    full_cat_path = " > ".join(selected_l1) + " > " + " > ".join(selected_l2) + " > " + " > ".join(selected_l3) + " > N/A"
                                else:
                                    col_l4_1, col_l4_2 = st.columns([3, 1])
                                    with col_l4_1:
                                        selected_l4 = st.multiselect("L4 Category (select one or multiple)", filtered_l4, key="cat_l4_multiselect")
                                    with col_l4_2:
                                        if st.button("Select All", key="select_all_l4", use_container_width=True):
                                            selected_l4 = filtered_l4
                                            st.session_state.cat_l4_multiselect = filtered_l4
                                            st.rerun()
                                        if st.button("Clear All", key="clear_all_l4", use_container_width=True):
                                            selected_l4 = []
                                            st.session_state.cat_l4_multiselect = []
                                            st.rerun()
                                    
                                    # Build full category path
                                    l1_str = ", ".join(selected_l1) if selected_l1 else "N/A"
                                    l2_str = ", ".join(selected_l2) if selected_l2 else "N/A"
                                    l3_str = ", ".join(selected_l3) if selected_l3 else "N/A"
                                    l4_str = ", ".join(selected_l4) if selected_l4 else "N/A"
                                    full_cat_path = f"{l1_str} > {l2_str} > {l3_str} > {l4_str}"
                            else:
                                selected_l4 = []
                                l1_str = ", ".join(selected_l1) if selected_l1 else "N/A"
                                l2_str = ", ".join(selected_l2) if selected_l2 else "N/A"
                                l3_str = ", ".join(selected_l3) if selected_l3 else "N/A"
                                full_cat_path = f"{l1_str} > {l2_str} > {l3_str} > N/A"
                    else:
                        selected_l3 = []
                        selected_l4 = []
                        l1_str = ", ".join(selected_l1) if selected_l1 else "N/A"
                        l2_str = ", ".join(selected_l2) if selected_l2 else "N/A"
                        full_cat_path = f"{l1_str} > {l2_str} > N/A > N/A"
            else:
                selected_l2 = []
                selected_l3 = []
                selected_l4 = []
                full_cat_path = "N/A > N/A > N/A > N/A"
    else:
        st.warning("No category data available")
        l1 = "N/A"
        l2 = "N/A"
        l3 = "N/A"
        l4 = "N/A"
        full_cat_path = "N/A > N/A > N/A > N/A"

# ==========================================
# 4. MAIN SCREEN
# ==========================================
st.title("🚦 Procurement Logic Capturer")

# Build context with multiple End Markets selection
context_parts = []
if 'selected_markets' in locals() and selected_markets:
    if len(selected_markets) == 1:
        context_parts.append(f"{selected_markets[0]} ({region})")
    elif len(selected_markets) <= 3:
        context_parts.append(f"{', '.join(selected_markets)} ({region})")
    else:
        context_parts.append(f"{len(selected_markets)} markets ({region})")
elif 'market' in locals() and market and market != "N/A":
    context_parts.append(f"{market} ({region})")
else:
    context_parts.append(f"({region})")

if 'business_user_markets' in locals() and business_user_markets:
    if len(business_user_markets) == 1:
        context_parts.append(f"Business User: {business_user_markets[0]}")
    elif len(business_user_markets) <= 3:
        context_parts.append(f"Business User: {', '.join(business_user_markets)}")
    else:
        context_parts.append(f"Business User: {len(business_user_markets)} markets")

if 'company_code' in locals() and company_code and company_code != "N/A":
    context_parts.append(f"Company: {company_code}")

# Handle multiple category selections
cat_display = full_cat_path if 'full_cat_path' in locals() else "N/A > N/A > N/A > N/A"
st.markdown(f"**Context:** {' | '.join(context_parts)} | **Category:** {cat_display}")

st.divider()

# ==========================================
# SUPPLIER TYPE SELECTION
# ==========================================
st.subheader("🏢 Supplier Type Selection")
supplier_type_filter = st.radio(
    "Filter by Supplier Type",
    options=["All", "Local", "Global"],
    horizontal=True,
    key="supplier_type_filter",
    help="Select whether to show Local, Global, or All suppliers in the pool"
)

st.divider()

# ==========================================
# SUPPLIER POOL TABLE (Above Streams)
# ==========================================
st.subheader("👥 Supplier Pool")

# Toggle to enable/disable supplier pool
enable_supplier_pool = st.toggle(
    "Enable Supplier Pool", 
    value=True, 
    key="enable_supplier_pool",
    help="When enabled, you can define a pool of suppliers. When disabled, you can switch to sourcing logic directly."
)

if enable_supplier_pool:
    st.info("Define suppliers with their logic type, supplier type, and buying channels. Each buying channel should be a separate row.")
    
    # Initialize suppliers_df in session state if not exists - start with completely empty row
    if 'suppliers_df' not in st.session_state or st.session_state.suppliers_df.empty:
        st.session_state.suppliers_df = pd.DataFrame([
            {
                "Supplier Name": "", 
                "Vendor Code": "", 
                "Supplier Type": "",
                "Logic Type": "",
                "Buying Channel": "",
                "Tender Required": "",
                "Comments": ""
            }
        ])
    
    # Ensure all columns exist and are of correct type
    required_columns = ["Supplier Name", "Vendor Code", "Supplier Type", "Logic Type", "Buying Channel", "Tender Required", "Comments"]
    for col in required_columns:
        if col not in st.session_state.suppliers_df.columns:
            st.session_state.suppliers_df[col] = ""
    
    # Convert to string type to ensure editability
    for col in ["Supplier Name", "Vendor Code", "Comments"]:
        if col in st.session_state.suppliers_df.columns:
            st.session_state.suppliers_df[col] = st.session_state.suppliers_df[col].astype(str)
    
    # Data Editor for Multiple Suppliers with enhanced fields
    column_config = {
        "Supplier Name": st.column_config.TextColumn(
            "Supplier Name", 
            required=False,
            default=""
        ),
        "Vendor Code": st.column_config.TextColumn(
            "Vendor Code (Optional)",
            default=""
        ),
        "Supplier Type": st.column_config.SelectboxColumn(
            "Supplier Type",
            options=["", "Local", "Global"],
            required=False,
            default=""
        ),
        "Logic Type": st.column_config.SelectboxColumn(
            "Logic Type",
            options=["", "Buying Channel", "Sourcing"],
            required=False,
            default="",
            help="Buying Channel: Direct purchase. Sourcing: Requires sourcing process."
        ),
        "Buying Channel": st.column_config.SelectboxColumn(
            "Buying Channel",
            options=["", "Hosted Catalog", "Punch-out", "Web Form", "Free Text", "P-Card"],
            required=False,
            default=""
        ),
        "Tender Required": st.column_config.SelectboxColumn(
            "Tender Required",
            options=["", "No", "Yes - Every Time", "Yes - Above Threshold"],
            required=False,
            default="",
            help="Whether this supplier requires a tender process"
        ),
        "Comments": st.column_config.TextColumn(
            "Comments / Logic Explanation",
            default=""
        ),
    }
    
    # Use the full dataframe for editing (don't filter before editing)
    suppliers_df = st.data_editor(
        st.session_state.suppliers_df,
        column_config=column_config,
        num_rows="dynamic", # Allow adding rows
        use_container_width=True,
        hide_index=True,
        key="suppliers_editor",
        disabled=False  # Ensure all cells are editable
    )
    
    # Update session state with edited data
    st.session_state.suppliers_df = suppliers_df.copy()
    
    # Filter by supplier type for display/visualization only (after saving)
    suppliers_df_filtered = suppliers_df.copy()
    if supplier_type_filter != "All":
        suppliers_df_filtered = suppliers_df_filtered[suppliers_df_filtered["Supplier Type"] == supplier_type_filter]
else:
    # Supplier pool disabled - create empty suppliers_df
    suppliers_df = pd.DataFrame(columns=["Supplier Name", "Vendor Code", "Supplier Type", "Logic Type", "Buying Channel", "Tender Required", "Comments"])

# Update supplier nodes for visualization (use filtered data if filter is applied)
supp_nodes_list = []
supp_node_ids = []
buying_channel_nodes = []
sourcing_nodes = []
local_suppliers = []
global_suppliers = []

# Use filtered suppliers for visualization based on supplier_type_filter
suppliers_for_viz = suppliers_df_filtered if 'suppliers_df_filtered' in locals() else suppliers_df

if suppliers_for_viz is not None and not suppliers_for_viz.empty:
    for idx, row in suppliers_for_viz.iterrows():
        supplier_name = str(row.get("Supplier Name", "")).strip()
        if supplier_name:
            node_id = f"Supp{idx}"
            supp_node_ids.append(node_id)
            # Sanitize text for Mermaid
            channel_clean = str(row.get("Buying Channel", "")).replace(":", "-").replace("<", "").replace(">", "").replace('"', "'")
            name_clean = supplier_name.replace(":", "-").replace("<", "").replace(">", "").replace('"', "'")
            supp_type = str(row.get("Supplier Type", "")).strip() or "Local"
            logic_type = str(row.get("Logic Type", "")).strip() or "Buying Channel"
            tender = str(row.get("Tender Required", "")).strip() or "No"
            supp_nodes_list.append(f'    {node_id}["{supp_type} {name_clean}\\n{channel_clean}\\n{logic_type}\\nTender: {tender}"]')
            
            # Categorize by supplier type and logic type
            if supp_type == "Local":
                local_suppliers.append(node_id)
            elif supp_type == "Global":
                global_suppliers.append(node_id)
            
            if logic_type == "Buying Channel":
                buying_channel_nodes.append(node_id)
            elif logic_type == "Sourcing":
                sourcing_nodes.append(node_id)

if not supp_nodes_list:
    supp_nodes_list = ['    NoSupp["No Defined Suppliers"]']
    supp_node_ids = ["NoSupp"]

st.divider()

# ==========================================
# TWO COLUMNS: Buying Channels & Sourcing Logic
# ==========================================
col_green, col_red = st.columns(2)

# ---------------------------------------------------------
# LEFT COLUMN: BUYING CHANNELS
# ---------------------------------------------------------
with col_green:
    st.markdown('<div class="header-style">⬅️ Buying Channels</div>', unsafe_allow_html=True)
    st.markdown('<div class="green-lane">', unsafe_allow_html=True)
    
    # Toggle for Buying Channels
    enable_buying_channels = st.toggle(
        "Enable Buying Channels", 
        value=True, 
        key="enable_buying_channels",
        help="When enabled, you can specify buying channels in the table below."
    )
    
    if enable_buying_channels:
        st.write("#### Buying Channels Configuration")
        st.info("Specify buying channels for suppliers. Each channel should be a separate row.")
        
        # Initialize buying_channels_df in session state if not exists - start with empty row
        if 'buying_channels_df' not in st.session_state:
            st.session_state.buying_channels_df = pd.DataFrame([
                {
                    "Channel Type": "",
                    "Supplier": "",
                    "Vendor Code": "",
                    "Link": "",
                    "Comments": ""
                }
            ])
        
        # Data Editor for Buying Channels - only Channel Type column
        buying_channel_config = {
            "Channel Type": st.column_config.SelectboxColumn(
                "Channel Type",
                options=["", "Hosted Catalog", "Punch-out", "Web Form", "Free Text", "P-Card"],
                required=False,
                default=""
            ),
            "Supplier": st.column_config.TextColumn("Supplier (Optional)", default=""),
            "Vendor Code": st.column_config.TextColumn("Vendor Code (Optional)", default=""),
            "Link": st.column_config.TextColumn("Link/URL (Optional)", default=""),
            "Comments": st.column_config.TextColumn("Comments", default=""),
        }
        
        buying_channels_df = st.data_editor(
            st.session_state.buying_channels_df,
            column_config=buying_channel_config,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            key="buying_channels_editor"
        )
        
        # Update session state
        st.session_state.buying_channels_df = buying_channels_df
        
        st.write("#### Marketplace Logic")
        allow_mkp = st.toggle("Allow Amazon / Marketplace?", key="allow_mkp_toggle")
        mkp_limit = 0
        if allow_mkp:
            mkp_limit = st.number_input("Auto-Approve Limit (£)", value=500, step=100, key="mkp_limit_input")
    else:
        # Buying channels disabled - create empty DataFrame
        buying_channels_df = pd.DataFrame(columns=["Channel Type", "Supplier", "Vendor Code", "Link", "Comments"])
        allow_mkp = False
        mkp_limit = 0
    
    # Set enable_stream1 based on buying channels toggle
    enable_stream1 = enable_buying_channels

    st.markdown('</div>', unsafe_allow_html=True)

# ---------------------------------------------------------
# RIGHT COLUMN: SOURCING LOGIC
# ---------------------------------------------------------
with col_red:
    st.markdown('<div class="header-style">➡️ Sourcing Logic</div>', unsafe_allow_html=True)
    st.markdown('<div class="red-lane">', unsafe_allow_html=True)
    
    # Toggle for Stream 2
    enable_stream2 = st.toggle("Enable Sourcing Logic", value=True, key="enable_stream2")
    
    if enable_stream2:
        st.write("#### Failover Logic")
        st.caption("If user cannot use Supplier Pool or Buying Channels, or if suppliers require sourcing:")
        
        threshold = st.number_input("⚡ Tactical vs Strategic Threshold (£)", value=10000, step=1000, key="threshold_input")
        
        st.markdown("---")
        st.markdown(f"**📉 Tactical (< £{threshold})**")
        tact_action = st.selectbox("Tactical Action", 
                                   ["Fairmarkit (Autonomous)", "3-Bids (Local Buyer)", "Spot Buy Desk", "No-Touch PO"],
                                   key="tact_action_select")
        
        st.markdown("---")
        st.markdown(f"**📈 Strategic (> £{threshold})**")
        strat_action = st.selectbox("Strategic Owner", 
                                    ["Global Category Lead", "Sourcing Manager", "Regional Hub", "RFP Team"],
                                    key="strat_action_select")
        
        st.markdown("---")
        instr = st.text_area("SDC / Desk Instructions", placeholder="e.g. Mandatory to use RFP template B...", key="instr_text_area")
    else:
        # Stream 2 disabled - set default values
        threshold = 10000
        tact_action = "N/A"
        strat_action = "N/A"
        instr = ""
    
    st.markdown('</div>', unsafe_allow_html=True)

# ==========================================
# RENDER LOGIC FLOW VISUALIZATION (After Stream 1 & 2)
# ==========================================
st.divider()
st.subheader("🗺️ Logic Flow Visualization")

# Check if any logic is enabled - only show visualization if at least one is enabled
enable_supplier_pool_val = enable_supplier_pool if 'enable_supplier_pool' in locals() else False
enable_buying_channels_val = enable_buying_channels if 'enable_buying_channels' in locals() else False
enable_stream2_val = enable_stream2 if 'enable_stream2' in locals() else False

# Only show visualization if at least one component is enabled
if not (enable_supplier_pool_val or enable_buying_channels_val or enable_stream2_val):
    st.warning("⚠️ Please enable at least one of: Supplier Pool, Buying Channels, or Sourcing Logic to see the visualization.")
else:
    # Show user context
    context_info = []
    if 'selected_markets' in locals() and selected_markets:
        context_info.append(f"**End Markets:** {', '.join(selected_markets[:3])}{'...' if len(selected_markets) > 3 else ''}")
    if 'business_user_markets' in locals() and business_user_markets:
        context_info.append(f"**Business User Markets:** {', '.join(business_user_markets[:3])}{'...' if len(business_user_markets) > 3 else ''}")
    if 'company_code' in locals() and company_code and company_code != "N/A":
        context_info.append(f"**Company Code:** {company_code}")
    if 'full_cat_path' in locals() and full_cat_path and full_cat_path != "N/A > N/A > N/A > N/A":
        context_info.append(f"**Category:** {full_cat_path}")
    
    # Show enabled components
    enabled_components = []
    if enable_supplier_pool_val:
        enabled_components.append("Supplier Pool")
    if enable_buying_channels_val:
        enabled_components.append("Buying Channels")
    if enable_stream2_val:
        enabled_components.append("Sourcing Logic")
    
    if enabled_components:
        context_info.append(f"**Enabled:** {', '.join(enabled_components)}")

    if context_info:
        st.info(" | ".join(context_info))

    # Prepare Supplier Nodes for Diagram (sanitized) - only if Supplier Pool is enabled
    supp_nodes_list = []
    supp_node_ids = []
    buying_channel_nodes = []
    sourcing_nodes = []
    local_suppliers = []
    global_suppliers = []

    if enable_supplier_pool_val and 'suppliers_df' in locals() and suppliers_df is not None and not suppliers_df.empty:
        for idx, row in suppliers_df.iterrows():
            supplier_name = str(row.get("Supplier Name", "")).strip()
            if supplier_name:
                node_id = f"Supp{idx}"
                supp_node_ids.append(node_id)
                # Sanitize text for Mermaid
                channel_clean = str(row.get("Buying Channel", "")).replace(":", "-").replace("<", "").replace(">", "").replace('"', "'")
                name_clean = supplier_name.replace(":", "-").replace("<", "").replace(">", "").replace('"', "'")
                supp_type = str(row.get("Supplier Type", "")).strip() or "Local"
                logic_type = str(row.get("Logic Type", "")).strip() or "Buying Channel"
                tender = str(row.get("Tender Required", "")).strip() or "No"
                
                # Create node label
                node_label = f"{name_clean}\\n{channel_clean}"
                if tender != "No":
                    node_label += f"\\n⚠️ Tender: {tender}"
                
                supp_nodes_list.append(f'    {node_id}["{node_label}"]')
                
                # Categorize by supplier type and logic type
                if supp_type == "Local":
                    local_suppliers.append(node_id)
                else:
                    global_suppliers.append(node_id)
                
                if logic_type == "Buying Channel":
                    buying_channel_nodes.append(node_id)
                else:
                    sourcing_nodes.append(node_id)

    if not supp_nodes_list and enable_supplier_pool_val:
        supp_nodes_list = ['    NoSupp["No Defined Suppliers"]']
        supp_node_ids = ["NoSupp"]

    # Get toggle states (use the actual values from the toggles)
    allow_mkp_val = allow_mkp if 'allow_mkp' in locals() and enable_buying_channels_val else False
    mkp_limit_val = mkp_limit if 'mkp_limit' in locals() and enable_buying_channels_val else 0
    threshold_val = threshold if 'threshold' in locals() and enable_stream2_val else 10000
    tact_action_val = tact_action if 'tact_action' in locals() and enable_stream2_val else "N/A"
    strat_action_val = strat_action if 'strat_action' in locals() and enable_stream2_val else "N/A"

    # Get category path for display
    cat_path_display = full_cat_path if 'full_cat_path' in locals() else "N/A"

    # Build Mermaid code with new flow: Taxonomy → Local/Global → Logic
    mermaid_lines = [
        "graph TD",
        "    Start([User Request]) --> CheckTaxonomy{Taxonomy Match?}",
    ]

    # Use full category path for display (no truncation)
    # Escape special characters for Mermaid
    cat_display_clean = cat_path_display.replace('"', "'").replace('\n', ' ')
    mermaid_lines.append(f'    CheckTaxonomy -->|Yes| CheckTaxonomyYes["Category: {cat_display_clean}"]')
    mermaid_lines.append("    CheckTaxonomy -->|No| Reject[Reject Request]")

    # Get supplier type filter
    supplier_filter_val = supplier_type_filter if 'supplier_type_filter' in locals() else "All"

    # Add supplier nodes first
    for node_line in supp_nodes_list:
        mermaid_lines.append(node_line)

    # Route based on supplier type filter selection - only if Supplier Pool is enabled
    if enable_supplier_pool_val:
        if supplier_filter_val == "Local":
            # Only show Local suppliers
            mermaid_lines.append("    CheckTaxonomyYes --> LocalPool((Local Pool))")
            if local_suppliers:
                if len(local_suppliers) == 1:
                    mermaid_lines.append(f"    LocalPool --> {local_suppliers[0]}")
                else:
                    mermaid_lines.append(f"    LocalPool --> {' --> '.join(local_suppliers)}")
            else:
                mermaid_lines.append("    LocalPool --> NoLocalSupp[No Local Suppliers]")
        elif supplier_filter_val == "Global":
            # Only show Global suppliers
            mermaid_lines.append("    CheckTaxonomyYes --> GlobalPool((Global Pool))")
            if global_suppliers:
                if len(global_suppliers) == 1:
                    mermaid_lines.append(f"    GlobalPool --> {global_suppliers[0]}")
                else:
                    mermaid_lines.append(f"    GlobalPool --> {' --> '.join(global_suppliers)}")
            else:
                mermaid_lines.append("    GlobalPool --> NoGlobalSupp[No Global Suppliers]")
        else:
            # Show both Local and Global (All)
            mermaid_lines.append("    CheckTaxonomyYes --> CheckSuppType{Local or Global Supplier?}")
            
            # Add supplier type branches
            if local_suppliers:
                mermaid_lines.append("    CheckSuppType -->|Local| LocalPool((Local Pool))")
                if len(local_suppliers) == 1:
                    mermaid_lines.append(f"    LocalPool --> {local_suppliers[0]}")
                else:
                    mermaid_lines.append(f"    LocalPool --> {' --> '.join(local_suppliers)}")
            
            if global_suppliers:
                mermaid_lines.append("    CheckSuppType -->|Global| GlobalPool((Global Pool))")
                if len(global_suppliers) == 1:
                    mermaid_lines.append(f"    GlobalPool --> {global_suppliers[0]}")
                else:
                    mermaid_lines.append(f"    GlobalPool --> {' --> '.join(global_suppliers)}")
            
            # If no suppliers match the filter, show message
            if not local_suppliers and not global_suppliers:
                mermaid_lines.append("    CheckSuppType -->|Any| NoSupp[No Suppliers Defined]")
    else:
        # Supplier Pool disabled - skip directly to next available logic
        mermaid_lines.append("    CheckTaxonomyYes --> CheckNextLogic{Next Logic?}")

    # Route suppliers based on logic type and toggles (for all filter types)
    # Only process if Supplier Pool is enabled
    if enable_supplier_pool_val and (local_suppliers or global_suppliers):
        if enable_buying_channels_val:
            # Buying Channels enabled - use buying channel logic
            if buying_channel_nodes:
                mermaid_lines.append("    %% BUYING CHANNEL ROUTE")
                for bc_node in buying_channel_nodes:
                    mermaid_lines.append(f"    {bc_node} --> BuyChannel[Use Buying Channel]")
            
            if sourcing_nodes:
                mermaid_lines.append("    %% SOURCING ROUTE")
                for src_node in sourcing_nodes:
                    if enable_stream2_val:
                        mermaid_lines.append(f"    {src_node} --> Sourcing")
                    else:
                        mermaid_lines.append(f"    {src_node} --> RejectSourcing[Reject - Sourcing Disabled]")
            
            # Failover from buying channels to sourcing
            if buying_channel_nodes and enable_stream2_val:
                mermaid_lines.append("    BuyChannel -.->|Failover| Sourcing")
        else:
            # Buying Channels disabled - suppliers go directly to sourcing or reject
            if buying_channel_nodes:
                mermaid_lines.append("    %% BUYING CHANNEL ROUTE (Buying Channels Disabled)")
                for bc_node in buying_channel_nodes:
                    # When Buying Channels is disabled, buying channel suppliers go to sourcing or reject
                    if enable_stream2_val:
                        mermaid_lines.append(f"    {bc_node} --> Sourcing")
                    else:
                        mermaid_lines.append(f"    {bc_node} --> RejectSourcing[Reject - Sourcing Disabled]")
            
            if sourcing_nodes:
                mermaid_lines.append("    %% SOURCING ROUTE")
                for src_node in sourcing_nodes:
                    if enable_stream2_val:
                        mermaid_lines.append(f"    {src_node} --> Sourcing")
                    else:
                        mermaid_lines.append(f"    {src_node} --> RejectSourcing[Reject - Sourcing Disabled]")
    elif enable_supplier_pool_val:
        # Supplier Pool enabled but no suppliers defined - check marketplace or sourcing
        if supplier_filter_val == "All":
            mermaid_lines.append("    CheckSuppType -->|Any| CheckSupp{Suppliers?}")
        else:
            mermaid_lines.append("    CheckTaxonomyYes --> CheckSupp{Suppliers?}")
    else:
        # Supplier Pool disabled - go directly to Buying Channels or Sourcing
        if enable_buying_channels_val:
            mermaid_lines.append("    CheckNextLogic -->|Buying Channels| BuyChannel[Use Buying Channel]")
        elif enable_stream2_val:
            mermaid_lines.append("    CheckNextLogic -->|Sourcing| Sourcing")
        else:
            mermaid_lines.append("    CheckNextLogic --> RejectAll[Reject - All Logic Disabled]")

    # Add marketplace logic (only if Buying Channels enabled and marketplace allowed)
    if enable_buying_channels_val and allow_mkp_val:
        mermaid_lines.extend([
            "",
            "    %% MARKETPLACE",
            "    CheckSupp -->|No| CheckMKP{Marketplace?}",
            f"    CheckMKP -->|Yes| MKPLimit{{< £{mkp_limit_val}?}}",
            "    MKPLimit -->|Yes| GoMKP[Buy on Marketplace]",
        ])
        if enable_stream2_val:
            mermaid_lines.append(f"    MKPLimit -->|No| Sourcing")
        else:
            mermaid_lines.append(f"    MKPLimit -->|No| RejectSourcing[Reject - Sourcing Disabled]")
        mermaid_lines.append("    CheckMKP -->|No| " + ("Sourcing" if enable_stream2_val else "RejectSourcing[Reject - Sourcing Disabled]"))
    elif not enable_buying_channels_val:
        # Buying Channels disabled - if no suppliers, go directly to Sourcing if enabled
        if enable_stream2_val:
            mermaid_lines.append("    CheckSupp -->|No| Sourcing")
        else:
            mermaid_lines.append("    CheckSupp -->|No| RejectAll[Reject - All Logic Disabled]")
    else:
        # Buying Channels enabled but no marketplace - failover to sourcing
        if enable_stream2_val:
            mermaid_lines.append("    CheckSupp -->|No| Sourcing")
        else:
            mermaid_lines.append("    CheckSupp -->|No| RejectSourcing[Reject - Sourcing Disabled]")

    # Add sourcing subgraph (only if Stream 2 enabled)
    if enable_stream2_val:
        mermaid_lines.extend([
            "",
            "    subgraph SourcingBox [Sourcing Logic]",
            "        direction TB",
            "        Sourcing(Start Sourcing) --> CheckThresh{> £" + str(threshold_val) + "?}",
            f'        CheckThresh -->|No| Tactical["Tactical: {tact_action_val}"]',
            f'        CheckThresh -->|Yes| Strategic["Strategic: {strat_action_val}"]',
            "    end",
        ])

    # Add styling
    mermaid_lines.extend([
        "",
        "    %% STYLING",
        "    classDef green fill:#dcfce7,stroke:#16a34a,stroke-width:2px",
        "    classDef red fill:#fee2e2,stroke:#ef4444,stroke-width:2px",
        "    classDef blue fill:#dbeafe,stroke:#3b82f6,stroke-width:2px",
        "    classDef yellow fill:#fef3c7,stroke:#f59e0b,stroke-width:2px",
    ])

    # Style nodes
    if supp_node_ids and supp_node_ids[0] != "NoSupp":
        mermaid_lines.append(f"    class {','.join(supp_node_ids)} green")
    if allow_mkp_val:
        mermaid_lines.append("    class GoMKP green")
    if enable_stream2_val:
        mermaid_lines.append("    class Tactical,Strategic red")
    mermaid_lines.append("    class Reject,RejectSourcing,RejectAll red")
    mermaid_lines.append("    class CheckTaxonomy,CheckSuppType blue")

    mermaid_code = "\n".join(mermaid_lines)

    # Display Mermaid code and download button
    col_viz1, col_viz2 = st.columns([3, 1])
    with col_viz1:
        with st.expander("📝 View Mermaid Code"):
            st.code(mermaid_code, language="text")

    with col_viz2:
        st.download_button(
            label="📥 Download Mermaid",
            data=mermaid_code,
            file_name=f"logic_flow_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.mmd",
            mime="text/plain",
            use_container_width=True
        )

    # Render Mermaid diagram
    components.html(f"""
    <div style="text-align:center; width:100%; padding:20px;">
        <div class="mermaid">{mermaid_code}</div>
    </div>
    <script src="https://cdn.jsdelivr.net/npm/mermaid@11/dist/mermaid.min.js"></script>
    <script>
        mermaid.initialize({{startOnLoad:true, theme:'default'}});
    </script>
    """, height=600, scrolling=True)

# ==========================================
# 5. FINAL OUTPUT (At the bottom)
# ==========================================
st.divider()
st.markdown("---")

# Initialize session state for output
if 'show_output' not in st.session_state:
    st.session_state.show_output = False

# Button to generate and show output
col1, col2, col3 = st.columns([1, 1, 2])
with col1:
    if st.button("📊 Generate Logic Output", type="primary", use_container_width=True):
        st.session_state.show_output = True
        st.rerun()

with col2:
    if st.button("🔄 Reset Output", use_container_width=True):
        st.session_state.show_output = False
        st.rerun()

if st.session_state.show_output:
    st.subheader("📋 Final Output - Ready for ORO Team")
    
    # Collect all data
    # Handle multiple category selections
    selected_l1_list = selected_l1 if 'selected_l1' in locals() and isinstance(selected_l1, list) else ([] if 'selected_l1' not in locals() else [selected_l1])
    selected_l2_list = selected_l2 if 'selected_l2' in locals() and isinstance(selected_l2, list) else ([] if 'selected_l2' not in locals() else [selected_l2])
    selected_l3_list = selected_l3 if 'selected_l3' in locals() and isinstance(selected_l3, list) else ([] if 'selected_l3' not in locals() else [selected_l3])
    selected_l4_list = selected_l4 if 'selected_l4' in locals() and isinstance(selected_l4, list) else ([] if 'selected_l4' not in locals() else [selected_l4])
    
    output_data = {
        "scope": {
            "region": region if 'region' in locals() else "N/A",
            "cluster": cluster if 'cluster' in locals() else "N/A",
            "end_markets": selected_markets if 'selected_markets' in locals() else [],
            "business_user_markets": business_user_markets if 'business_user_markets' in locals() else [],
            "company_code": company_code if 'company_code' in locals() else "N/A"
        },
        "category": {
            "full_path": full_cat_path if 'full_cat_path' in locals() else "N/A",
            "l1": selected_l1_list,
            "l2": selected_l2_list,
            "l3": selected_l3_list,
            "l4": selected_l4_list
        },
        "supplier_pool": {
            "enabled": enable_supplier_pool if 'enable_supplier_pool' in locals() else True,
            "suppliers": suppliers_df.to_dict('records') if 'suppliers_df' in locals() and suppliers_df is not None and not suppliers_df.empty and len(suppliers_df) > 0 and (suppliers_df.iloc[0].get("Supplier Name", "") if len(suppliers_df) > 0 else "") else [],
            "supplier_type_filter": supplier_type_filter if 'supplier_type_filter' in locals() else "All"
        },
        "buying_channels": {
            "enabled": enable_buying_channels if 'enable_buying_channels' in locals() else True,
            "channels": buying_channels_df.to_dict('records') if 'buying_channels_df' in locals() and buying_channels_df is not None and not buying_channels_df.empty else [],
            "allow_marketplace": allow_mkp if 'allow_mkp' in locals() else False,
            "marketplace_limit": mkp_limit if 'mkp_limit' in locals() else 0
        },
        "stream2": {
            "enabled": enable_stream2 if 'enable_stream2' in locals() else True,
            "tactical_threshold": threshold if 'threshold' in locals() else 0,
            "tactical_action": tact_action if 'tact_action' in locals() else "N/A",
            "strategic_owner": strat_action if 'strat_action' in locals() else "N/A",
            "instructions": instr if 'instr' in locals() else ""
        },
        "metadata": {
            "created_at": pd.Timestamp.now().isoformat(),
            "version": "2.0"
        }
    }
    
    # Display JSON Blueprint
    st.markdown("### 📄 JSON Blueprint")
    json_output = json.dumps(output_data, indent=2, ensure_ascii=False)
    st.code(json_output, language="json")
    
    # Download buttons
    col_dl1, col_dl2, col_dl3 = st.columns(3)
    
    with col_dl1:
        st.download_button(
            label="💾 Download JSON",
            data=json_output,
            file_name=f"oro_logic_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.json",
            mime="application/json",
            use_container_width=True
        )
    
    with col_dl2:
        # Create Excel file
        if OPENPYXL_AVAILABLE:
            try:
                import io
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                
                wb = Workbook()
                
                # Sheet 1: Logic Matrix
                ws1 = wb.active
                ws1.title = "Logic Matrix"
                ws1.append(["Field", "Value"])
                ws1.append(["Region", output_data["scope"]["region"]])
                ws1.append(["Cluster/DRBU", output_data["scope"]["cluster"]])
                ws1.append(["End Markets", ", ".join(output_data["scope"]["end_markets"]) if output_data["scope"]["end_markets"] else "N/A"])
                ws1.append(["Business User Markets", ", ".join(output_data["scope"]["business_user_markets"]) if output_data["scope"]["business_user_markets"] else "N/A"])
                ws1.append(["Company Code", output_data["scope"]["company_code"]])
                ws1.append(["Category L1", ", ".join(output_data["category"]["l1"]) if output_data["category"]["l1"] else "N/A"])
                ws1.append(["Category L2", ", ".join(output_data["category"]["l2"]) if output_data["category"]["l2"] else "N/A"])
                ws1.append(["Category L3", ", ".join(output_data["category"]["l3"]) if output_data["category"]["l3"] else "N/A"])
                ws1.append(["Category L4", ", ".join(output_data["category"]["l4"]) if output_data["category"]["l4"] else "N/A"])
                ws1.append(["Category Full Path", output_data["category"]["full_path"]])
                ws1.append(["Supplier Pool Enabled", output_data["supplier_pool"]["enabled"]])
                ws1.append(["Supplier Type Filter", output_data["supplier_pool"]["supplier_type_filter"]])
                ws1.append(["Buying Channels Enabled", output_data["buying_channels"]["enabled"]])
                ws1.append(["Allow Marketplace", output_data["buying_channels"]["allow_marketplace"]])
                ws1.append(["Marketplace Limit", output_data["buying_channels"]["marketplace_limit"]])
                ws1.append(["Stream 2 Enabled", output_data["stream2"]["enabled"]])
                ws1.append(["Tactical Threshold", output_data["stream2"]["tactical_threshold"]])
                ws1.append(["Tactical Action", output_data["stream2"]["tactical_action"]])
                ws1.append(["Strategic Owner", output_data["stream2"]["strategic_owner"]])
                
                # Sheet 2: Suppliers
                ws2 = wb.create_sheet("Suppliers")
                if output_data["supplier_pool"]["suppliers"]:
                    ws2.append(["Supplier Name", "Vendor Code", "Supplier Type", "Logic Type", "Buying Channel", "Tender Required", "Comments"])
                    for supp in output_data["supplier_pool"]["suppliers"]:
                        ws2.append([
                            supp.get("Supplier Name", ""),
                            supp.get("Vendor Code", ""),
                            supp.get("Supplier Type", ""),
                            supp.get("Logic Type", ""),
                            supp.get("Buying Channel", ""),
                            supp.get("Tender Required", ""),
                            supp.get("Comments", "")
                        ])
                else:
                    ws2.append(["No suppliers defined"])
                
                # Sheet 3: Buying Channels
                ws3_bc = wb.create_sheet("Buying Channels")
                if output_data["buying_channels"]["channels"]:
                    ws3_bc.append(["Channel Type", "Supplier", "Vendor Code", "Link", "Comments"])
                    for ch in output_data["buying_channels"]["channels"]:
                        ws3_bc.append([
                            ch.get("Channel Type", ""),
                            ch.get("Supplier", ""),
                            ch.get("Vendor Code", ""),
                            ch.get("Link", ""),
                            ch.get("Comments", "")
                        ])
                else:
                    ws3_bc.append(["No buying channels defined"])
                
                # Sheet 4: Summary
                ws4 = wb.create_sheet("Summary")
                ws4.append(["Item", "Count"])
                ws4.append(["End Markets", len(output_data["scope"]["end_markets"])])
                ws4.append(["Business User Markets", len(output_data["scope"]["business_user_markets"])])
                ws4.append(["Suppliers", len(output_data["supplier_pool"]["suppliers"])])
                ws4.append(["Buying Channels", len(output_data["buying_channels"]["channels"])])
                
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                st.download_button(
                    label="📊 Download Excel",
                    data=excel_buffer.getvalue(),
                    file_name=f"oro_logic_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"Excel export error: {str(e)}")
        else:
            st.download_button(
                label="📊 Download Excel",
                data="",  # Empty data since openpyxl is not available
                file_name="",
                disabled=True,
                use_container_width=True,
                help="Install openpyxl: pip install openpyxl"
            )
            st.info("💡 Install openpyxl to enable Excel export: `pip install openpyxl`")
    
    with col_dl3:
        # Copy to clipboard button (JSON)
        if st.button("📋 Copy JSON to Clipboard", use_container_width=True):
            st.code(json_output, language="json")
            st.success("JSON copied! (Use Ctrl+C to copy from the code block above)")
    
    # Share section
    st.markdown("---")
    st.markdown("### 🔗 Share with ORO Team")
    st.info("💡 **Share Options:**")
    st.markdown("""
    1. **Download JSON** - Share the JSON file with the ORO team for integration
    2. **Download Excel** - Share the Excel file for review and validation
    3. **Copy Mermaid Code** - Share the Mermaid code for diagram visualization
    4. **Screenshot** - Take a screenshot of the visualization above
    """)
    
    # Summary table
    st.markdown("### 📊 Summary")
    summary_df = pd.DataFrame([
        ["Scope", f"{output_data['scope']['region']} / {output_data['scope']['cluster']}"],
        ["End Markets", len(output_data['scope']['end_markets'])],
        ["Business User Markets", len(output_data['scope']['business_user_markets'])],
        ["Category", output_data['category']['full_path']],
        ["Suppliers", len(output_data['supplier_pool']['suppliers'])],
        ["Buying Channels", len(output_data['buying_channels']['channels'])],
        ["Marketplace Enabled", "Yes" if output_data['buying_channels']['allow_marketplace'] else "No"],
        ["Tactical Threshold", f"£{output_data['stream2']['tactical_threshold']}"],
    ], columns=["Item", "Value"])
    st.dataframe(summary_df, use_container_width=True, hide_index=True)
else:
    st.info("👆 Click 'Generate Logic Output' to create and view the final output")